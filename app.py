import pandas as pd
from fastapi import FastAPI, File, UploadFile, HTTPException
from typing import Dict, List, Optional, Tuple
import io
import openpyxl
from openpyxl.utils import get_column_letter
import xlrd
from pydantic import BaseModel
import hashlib
from fastapi.responses import JSONResponse

app = FastAPI()

# ========== DATA MODELS ==========
class TableRow(BaseModel):
    name: str
    values: List[Optional[float]]
    location: str

class TableData(BaseModel):
    name: str
    sheet: str
    start_row: int
    end_row: int
    start_col: str
    end_col: str
    rows: Dict[str, TableRow]  # row_name: TableRow

class UploadedFile(BaseModel):
    filename: str
    content_hash: str
    sheets: List[str]
    tables: Dict[str, TableData]  # table_name: TableData

# ========== GLOBAL STORE ==========
file_store: Dict[str, UploadedFile] = {}

# ========== UTILITIES ==========
def get_file_hash(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()

def convert_to_float(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        value = value.strip()
        if '%' in value:
            try:
                return float(value.replace('%', '')) / 100
            except ValueError:
                return None
        if '$' in value:
            try:
                return float(value.replace('$', '').replace(',', ''))
            except ValueError:
                return None
        try:
            return float(value)
        except ValueError:
            return None
    return None

# ========== EXCEL PROCESSORS ==========
async def process_xlsx_file(content: bytes, filename: str) -> UploadedFile:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheets = wb.sheetnames
    tables = {}

    for sheet_name in sheets:
        sheet = wb[sheet_name]
        
        # Scan for tables with ALL CAPS headers
        for row in sheet.iter_rows():
            for cell in row:
                if (cell.value and isinstance(cell.value, str) and 
                    cell.value.isupper() and cell.value.strip() == cell.value):
                    
                    # Verify it's a valid table header
                    if not is_valid_table_header(sheet, cell):
                        continue
                    
                    table_name = cell.value.strip()
                    start_row = cell.row
                    start_col = cell.column
                    
                    # Find table boundaries
                    data_start_row, end_row, end_col = find_table_boundaries(sheet, start_row, start_col)
                    if not data_start_row:
                        continue
                    
                    # Extract all row data
                    rows = {}
                    for row_idx in range(data_start_row, end_row + 1):
                        row_cells = [
                            sheet.cell(row=row_idx, column=col_idx).value 
                            for col_idx in range(start_col, end_col + 1)
                        ]
                        if not row_cells or not row_cells[0]:
                            continue
                            
                        row_name = str(row_cells[0]).strip()
                        values = [convert_to_float(v) for v in row_cells[1:]]
                        location = f"{get_column_letter(start_col)}{row_idx}"
                        
                        rows[row_name] = TableRow(
                            name=row_name,
                            values=values,
                            location=location
                        )
                    
                    if rows:
                        tables[table_name] = TableData(
                            name=table_name,
                            sheet=sheet_name,
                            start_row=data_start_row,
                            end_row=end_row,
                            start_col=get_column_letter(start_col),
                            end_col=get_column_letter(end_col),
                            rows=rows
                        )
    
    return UploadedFile(
        filename=filename,
        content_hash=get_file_hash(content),
        sheets=sheets,
        tables=tables
    )

async def process_xls_file(content: bytes, filename: str) -> UploadedFile:
    file_like = io.BytesIO(content)
    wb = xlrd.open_workbook(file_contents=content)
    sheets = wb.sheet_names()
    tables = {}

    for sheet_name in sheets:
        sheet = wb.sheet_by_name(sheet_name)
        
        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                if (cell.value and isinstance(cell.value, str) and 
                    cell.value.isupper() and cell.value.strip() == cell.value):
                    
                    # Verify it's a valid table header
                    if not is_valid_xls_table_header(sheet, row_idx, col_idx):
                        continue
                    
                    table_name = cell.value.strip()
                    start_row = row_idx
                    start_col = col_idx
                    
                    # Find table boundaries
                    data_start_row, end_row, end_col = find_xls_table_boundaries(sheet, start_row, start_col)
                    if not data_start_row:
                        continue
                    
                    # Extract all row data
                    rows = {}
                    for data_row_idx in range(data_start_row, end_row + 1):
                        row_cells = [
                            sheet.cell_value(data_row_idx, c) 
                            for c in range(start_col, end_col + 1)
                        ]
                        if not row_cells or not row_cells[0]:
                            continue
                            
                        row_name = str(row_cells[0]).strip()
                        values = [convert_to_float(v) for v in row_cells[1:]]
                        location = f"{get_column_letter(start_col + 1)}{data_row_idx + 1}"
                        
                        rows[row_name] = TableRow(
                            name=row_name,
                            values=values,
                            location=location
                        )
                    
                    if rows:
                        tables[table_name] = TableData(
                            name=table_name,
                            sheet=sheet_name,
                            start_row=data_start_row + 1,
                            end_row=end_row + 1,
                            start_col=get_column_letter(start_col + 1),
                            end_col=get_column_letter(end_col + 1),
                            rows=rows
                        )
    
    return UploadedFile(
        filename=filename,
        content_hash=get_file_hash(content),
        sheets=sheets,
        tables=tables
    )

# ========== TABLE DETECTION HELPERS ==========
def is_valid_table_header(sheet, cell) -> bool:
    """Check if a cell is a valid table header in xlsx"""
    if cell.column > 1 and sheet.cell(row=cell.row, column=cell.column-1).value:
        return False
    
    empty_adjacent = 0
    for col in range(cell.column + 1, min(cell.column + 4, sheet.max_column + 1)):
        if not sheet.cell(row=cell.row, column=col).value:
            empty_adjacent += 1
        else:
            break
    
    if empty_adjacent < 2:
        return False
    
    if cell.row + 1 > sheet.max_row:
        return False
        
    next_row_has_data = False
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=cell.row + 1, column=col).value:
            next_row_has_data = True
            break
    
    return next_row_has_data

def is_valid_xls_table_header(sheet, row_idx, col_idx) -> bool:
    """Check if a cell is a valid table header in xls"""
    if col_idx > 0 and sheet.cell_value(row_idx, col_idx-1):
        return False
    
    empty_adjacent = 0
    for c in range(col_idx + 1, min(col_idx + 4, sheet.ncols)):
        if not sheet.cell_value(row_idx, c):
            empty_adjacent += 1
        else:
            break
    
    if empty_adjacent < 2:
        return False
    
    if row_idx + 1 >= sheet.nrows:
        return False
        
    next_row_has_data = False
    for c in range(sheet.ncols):
        if sheet.cell_value(row_idx + 1, c):
            next_row_has_data = True
            break
    
    return next_row_has_data

def find_table_boundaries(sheet, start_row, start_col) -> Tuple[int, int, int]:
    """Find table boundaries in xlsx"""
    data_start_row = start_row + 1
    while data_start_row <= sheet.max_row:
        next_row_empty = True
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=data_start_row, column=col).value:
                next_row_empty = False
                break
        if not next_row_empty:
            break
        data_start_row += 1
    
    if data_start_row > sheet.max_row:
        return (None, None, None)
    
    end_row = data_start_row
    while end_row <= sheet.max_row:
        has_header = False
        for col in range(1, sheet.max_column + 1):
            cell_val = sheet.cell(row=end_row, column=col).value
            if (cell_val and isinstance(cell_val, str) and 
                cell_val.isupper() and cell_val.strip() == cell_val and
                is_valid_table_header(sheet, sheet.cell(row=end_row, column=col))):
                has_header = True
                break
        if has_header:
            break
        
        row_empty = True
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=end_row, column=col).value:
                row_empty = False
                break
        if row_empty:
            break
        
        end_row += 1
    
    end_row = min(end_row - 1, sheet.max_row)
    
    end_col = start_col
    for col in range(start_col, sheet.max_column + 1):
        col_has_data = False
        for row_num in range(data_start_row, end_row + 1):
            if sheet.cell(row=row_num, column=col).value:
                col_has_data = True
                break
        if not col_has_data:
            break
        end_col = col
    
    return (data_start_row, end_row, end_col)

def find_xls_table_boundaries(sheet, start_row, start_col) -> Tuple[int, int, int]:
    """Find table boundaries in xls"""
    data_start_row = start_row + 1
    while data_start_row < sheet.nrows:
        next_row_empty = True
        for c in range(sheet.ncols):
            if sheet.cell_value(data_start_row, c):
                next_row_empty = False
                break
        if not next_row_empty:
            break
        data_start_row += 1
    
    if data_start_row >= sheet.nrows:
        return (None, None, None)
    
    end_row = data_start_row
    while end_row < sheet.nrows:
        has_header = False
        for c in range(sheet.ncols):
            cell_val = sheet.cell_value(end_row, c)
            if (cell_val and isinstance(cell_val, str) and 
                cell_val.isupper() and cell_val.strip() == cell_val and
                is_valid_xls_table_header(sheet, end_row, c)):
                has_header = True
                break
        if has_header:
            break
        
        row_empty = True
        for c in range(sheet.ncols):
            if sheet.cell_value(end_row, c):
                row_empty = False
                break
        if row_empty:
            break
        
        end_row += 1
    
    end_row = min(end_row - 1, sheet.nrows - 1)
    
    end_col = start_col
    for c in range(start_col, sheet.ncols):
        col_has_data = False
        for row_num in range(data_start_row, end_row + 1):
            if sheet.cell_value(row_num, c):
                col_has_data = True
                break
        if not col_has_data:
            break
        end_col = c
    
    return (data_start_row, end_row, end_col)

# ========== API ENDPOINTS ==========
@app.post("/uploadfile/")
async def upload_file(file: UploadFile):
    try:
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="Empty file")
        
        if not file.filename.endswith(('.xls', '.xlsx')):
            raise HTTPException(status_code=400, detail="Only Excel files allowed")
        
        file_hash = get_file_hash(content)
        
        if file_hash in file_store:
            return JSONResponse(
                status_code=200,
                content={
                    "file_id": file_hash,
                    "message": "File already processed",
                    "tables": list(file_store[file_hash].tables.keys())
                }
            )
        
        try:
            if file.filename.endswith('.xlsx'):
                uploaded_file = await process_xlsx_file(content, file.filename)
            else:
                uploaded_file = await process_xls_file(content, file.filename)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")
        
        file_store[file_hash] = uploaded_file
        
        return {
            "file_id": file_hash,
            "filename": uploaded_file.filename,
            "sheets": uploaded_file.sheets,
            "tables": list(uploaded_file.tables.keys()),
            "message": "File processed successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Server error: {str(e)}")

@app.get("/list_tables")
async def list_tables(file_id: str):
    if file_id not in file_store:
        raise HTTPException(status_code=404, detail="File not found")
    
    uploaded_file = file_store[file_id]
    return {
        "tables": [
            {
                "name": table.name,
                "sheet": table.sheet,
                "location": f"{table.start_col}{table.start_row}:{table.end_col}{table.end_row}",
                "row_count": len(table.rows)
            }
            for table in uploaded_file.tables.values()
        ]
    }

@app.get("/get_table_details")
async def get_table_details(file_id: str, table_name: str):
    if file_id not in file_store:
        raise HTTPException(status_code=404, detail="File not found")
    
    uploaded_file = file_store[file_id]
    table = uploaded_file.tables.get(table_name)
    
    if not table:
        raise HTTPException(status_code=404, detail="Table not found")
    
    return {
        "table_name": table.name,
        "sheet": table.sheet,
        "row_names": list(table.rows.keys()),
        "location": f"{table.start_col}{table.start_row}:{table.end_col}{table.end_row}"
    }

@app.get("/row_value")
async def get_row_value(file_id: str, table_name: str, row_name: str):
    if file_id not in file_store:
        raise HTTPException(status_code=404, detail="File not found")
    
    uploaded_file = file_store[file_id]
    table = uploaded_file.tables.get(table_name)
    
    if not table:
        raise HTTPException(status_code=404, detail="Table not found")
    
    row = table.rows.get(row_name)
    if not row:
        raise HTTPException(status_code=404, detail="Row not found")
    
    if not row.values:
        return {
            "table_name": table_name,
            "row_name": row_name,
            "value": None,
            "sheet": table.sheet,
            "location": row.location
        }
    
    # For single value rows, return the first value
    if len(row.values) == 1:
        value = row.values[0]
    else:
        # For multiple values, return the sum (ignoring None values)
        value = sum(v for v in row.values if v is not None)
    
    return {
        "table_name": table_name,
        "row_name": row_name,
        "value": value,
        "sheet": table.sheet,
        "location": row.location
    }